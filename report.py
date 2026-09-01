"""Report generation: CSV and a formatted multi-sheet Excel workbook."""

from __future__ import annotations

import os
from datetime import datetime
from typing import Dict, List

import pandas as pd

from .analyzer import (
    PageResult,
    build_summary,
    format_issue_rows,
    health_score,
    issue_breakdown,
)
from .utils import truncate

COLUMNS = [
    "URL",
    "Final URL",
    "Status Code",
    "Response Time (s)",
    "Content Type",
    "Depth",
    "Indexability",
    "Indexability Reason",
    "Title",
    "Title Length",
    "Duplicate Title",
    "Meta Description",
    "Meta Length",
    "Duplicate Meta",
    "H1",
    "H1 Count",
    "H2 Count",
    "Canonical",
    "Canonical Status",
    "Meta Robots",
    "X-Robots-Tag",
    "Lang",
    "Hreflang Tags",
    "Schema Types",
    "Internal Links",
    "External Links",
    "Nofollow Links",
    "Images",
    "Images Missing ALT",
    "Word Count",
    "Issue Count",
    "Issues",
]


def results_to_dataframe(results: List[PageResult]) -> pd.DataFrame:
    """Flatten crawl results into the main report table."""
    rows: List[Dict[str, object]] = []
    for r in results:
        seo = r.seo
        rows.append(
            {
                "URL": r.url,
                "Final URL": r.final_url,
                "Status Code": r.status_code if r.status_code is not None else "",
                "Response Time (s)": r.response_time,
                "Content Type": (r.content_type or "").split(";")[0],
                "Depth": r.depth,
                "Indexability": r.indexability,
                "Indexability Reason": r.indexability_reason,
                "Title": seo.title,
                "Title Length": seo.title_length,
                "Duplicate Title": "Yes" if r.duplicate_title else "No",
                "Meta Description": seo.meta_description,
                "Meta Length": seo.meta_description_length,
                "Duplicate Meta": "Yes" if r.duplicate_meta_description else "No",
                "H1": seo.h1,
                "H1 Count": seo.h1_count,
                "H2 Count": len(seo.h2_list),
                "Canonical": seo.canonical,
                "Canonical Status": r.canonical_status,
                "Meta Robots": seo.meta_robots,
                "X-Robots-Tag": r.x_robots_tag,
                "Lang": seo.lang,
                "Hreflang Tags": seo.hreflang_count,
                "Schema Types": ", ".join(seo.schema_types),
                "Internal Links": len(seo.internal_links),
                "External Links": len(seo.external_links),
                "Nofollow Links": seo.nofollow_links,
                "Images": seo.images_total,
                "Images Missing ALT": seo.images_missing_alt,
                "Word Count": seo.word_count,
                "Issue Count": r.issue_count,
                "Issues": truncate(" | ".join(r.issues), 500),
            }
        )
    return pd.DataFrame(rows, columns=COLUMNS)


def summary_to_dataframe(results: List[PageResult], start_url: str) -> pd.DataFrame:
    """Key/value table used for the Summary sheet."""
    summary = build_summary(results)
    rows = [
        {"Metric": "Website", "Value": start_url},
        {"Metric": "Crawl date", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Metric": "SEO health score (0-100)", "Value": health_score(results)},
    ]
    rows += [{"Metric": key, "Value": value} for key, value in summary.items()]
    return pd.DataFrame(rows)


def issues_to_dataframe(results: List[PageResult]) -> pd.DataFrame:
    rows = format_issue_rows(results)
    if not rows:
        return pd.DataFrame(columns=["URL", "Severity", "Issue", "Status Code", "Indexability", "Title"])
    return pd.DataFrame(rows)


def breakdown_to_dataframe(results: List[PageResult]) -> pd.DataFrame:
    breakdown = issue_breakdown(results)
    total = len(results) or 1
    return pd.DataFrame(
        [
            {
                "Issue Type": name,
                "Pages Affected": count,
                "% of Crawled Pages": round(count / total * 100, 1),
            }
            for name, count in breakdown.items()
        ]
    )


# ----------------------------------------------------------------------
# Writers
# ----------------------------------------------------------------------

def _autosize(worksheet, dataframe: pd.DataFrame, max_width: int = 60) -> None:
    from openpyxl.utils import get_column_letter

    for index, column in enumerate(dataframe.columns, start=1):
        series = dataframe[column].astype(str)
        longest = max([len(str(column))] + [len(v) for v in series.head(200)]) + 2
        worksheet.column_dimensions[get_column_letter(index)].width = min(longest, max_width)


def _style_header(worksheet, columns_count: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1][:columns_count]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="left")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def write_csv(results: List[PageResult], path: str) -> str:
    df = results_to_dataframe(results)
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def write_issues_csv(results: List[PageResult], path: str) -> str:
    df = issues_to_dataframe(results)
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def write_excel(results: List[PageResult], path: str, start_url: str) -> str:
    """Write a four-sheet workbook: Summary, Crawl Data, Issues, Issue Breakdown."""
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)

    sheets = {
        "Summary": summary_to_dataframe(results, start_url),
        "Crawl Data": results_to_dataframe(results),
        "Issues": issues_to_dataframe(results),
        "Issue Breakdown": breakdown_to_dataframe(results),
    }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
            worksheet = writer.sheets[name]
            if not frame.empty:
                _style_header(worksheet, len(frame.columns))
                _autosize(worksheet, frame)
    return path


def generate_reports(
    results: List[PageResult],
    start_url: str,
    output_dir: str = "output",
    fmt: str = "both",
    prefix: str = "",
) -> List[str]:
    """Write the requested report formats and return the file paths."""
    from .utils import get_host, safe_filename

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = safe_filename(prefix) if prefix else f"seo_crawl_{safe_filename(get_host(start_url))}_{stamp}"
    written: List[str] = []

    if fmt in ("csv", "both"):
        written.append(write_csv(results, os.path.join(output_dir, f"{base}.csv")))
        written.append(
            write_issues_csv(results, os.path.join(output_dir, f"{base}_issues.csv"))
        )
    if fmt in ("xlsx", "excel", "both"):
        written.append(
            write_excel(results, os.path.join(output_dir, f"{base}.xlsx"), start_url)
        )
    return written
